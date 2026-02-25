import streamlit as st
import pandas as pd
from datetime import datetime

from config_loader import config
from db_connector import db
from data_loader import (
    load_shop_config_from_db,
    get_available_months_incentives,
    load_incentives_by_employee,
    load_incentives_by_month_employee,
    load_incentives_details,
    create_monthly_pivot_table,
    calculate_incentive_summary_metrics
)

from charts import (
    create_incentive_pie_chart,
    create_incentive_bar_chart,
    create_monthly_stores_comparison_chart, 
    format_incentive_monthly_table
)

settings = config.get_settings()
ui_config = settings.get('ui', {})

st.set_page_config(
    page_title="Incentivos - PowerX",
    page_icon="💰",
    layout=ui_config.get('layout', 'wide')
)

def format_brl(valor):
    """Formata valor em Real Brasileiro"""
    if pd.isna(valor) or valor == 0:
        return "R$ 0,00"
    return f"R$ {valor:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')


def extrair_funcoes_unicas(df_employee: pd.DataFrame) -> list:
    funcoes = set()
    for valor in df_employee['funcao'].dropna():
        for f in valor.split(' / '):
            f = f.strip()
            if f and f != 'Não informado':
                funcoes.add(f)
    return sorted(list(funcoes))


def filtrar_por_funcao(df_employee: pd.DataFrame, funcoes_selecionadas: list) -> pd.DataFrame:
    if not funcoes_selecionadas:
        return df_employee
    
    mask = df_employee['funcao'].apply(
        lambda x: any(
            f.strip() in funcoes_selecionadas
            for f in str(x).split(' / ')
        ) if pd.notna(x) else False
    )
    return df_employee[mask]

if not db.test_connection():
    st.error("❌ Não foi possível conectar ao PostgreSQL")
    st.stop()

col_nav1, col_nav2 = st.columns([1, 6])
with col_nav1:
    if st.button("← Voltar para Vendas", use_container_width=True):
        st.switch_page("streamlit_app.py")

st.title('Análise de Incentivos - Dashboard PowerX')
st.markdown('---')

st.sidebar.header('Filtros de Incentivos')

shop_config = load_shop_config_from_db()

if not shop_config:
    st.error("Não achou dados no banco de dados")
    st.stop()

grupos = list(shop_config.keys())

if not grupos:
    st.error("Nenhum grupo encontrado")
    st.stop()

grupo_selecionado = st.sidebar.selectbox(
    '🏢 Selecione o Grupo',
    options=grupos,
    index=0
)

lojas_do_grupo = list(shop_config[grupo_selecionado].keys())

lojas_selecionadas = st.sidebar.multiselect(
    'Selecione Loja(s)',
    options=lojas_do_grupo,
    default=lojas_do_grupo
)

if not lojas_selecionadas:
    st.warning("Escolha pelo menos 1 loja")
    st.stop()

lojas_configs = [
    {
        'id': shop_config[grupo_selecionado][loja]['id'],
        'nome': loja,
        'cnpj': shop_config[grupo_selecionado][loja].get('cnpj', '')
    }
    for loja in lojas_selecionadas
]

lojas_ids = [loja['id'] for loja in lojas_configs]

st.sidebar.markdown('---')
st.sidebar.subheader('Período de Análise')

meses_disponiveis = get_available_months_incentives()

if not meses_disponiveis:
    st.error("Nenhum mês com dados encontrado")
    st.stop()

meses_options = {
    mes['mes_display']: mes['mes'] 
    for mes in meses_disponiveis
}

default_meses_display = list(meses_options.keys())[:3]

meses_selecionados_display = st.sidebar.multiselect(
    'Selecione os Meses',
    options=list(meses_options.keys()),
    default=default_meses_display,
    help="Selecione um ou mais meses para análise"
)

if not meses_selecionados_display:
    st.warning("Selecione pelo menos um mês")
    st.stop()

meses_selecionados = [meses_options[m] for m in meses_selecionados_display]

st.sidebar.markdown('---')
st.sidebar.subheader('Configurações')

top_n_vendedores = st.sidebar.slider(
    'Top N Premiados nos Gráficos',
    min_value=5,
    max_value=20,
    value=10,
    step=1
)

with st.spinner('Carregando dados de incentivos...'):
    df_employee = load_incentives_by_employee(lojas_ids, meses_selecionados)
    df_monthly  = load_incentives_by_month_employee(lojas_ids, meses_selecionados)
    df_details  = load_incentives_details(lojas_ids, meses_selecionados)

if df_employee.empty:
    titulo_lojas = ', '.join(lojas_selecionadas) if len(lojas_selecionadas) < 4 else f"{len(lojas_selecionadas)} Lojas"
    st.warning(
        f"**Sem dados de incentivos** para **{titulo_lojas}** nos meses selecionados: "
        f"**{', '.join(meses_selecionados_display)}**"
    )
    st.info("Tente selecionar outros meses ou outras lojas")
    st.stop()


metrics      = calculate_incentive_summary_metrics(df_employee)
titulo_lojas = ', '.join(lojas_selecionadas) if len(lojas_selecionadas) < 4 else f"{len(lojas_selecionadas)} Lojas"

st.subheader(f'Grupo: {grupo_selecionado} | Lojas: {titulo_lojas}')
st.caption(f'Período: {", ".join(meses_selecionados_display)}')

st.markdown('---')

col1, col2, col3 = st.columns(3)

col1.metric("Valor Total",  format_brl(metrics['valor_total']))
col2.metric("Premiados",    f"{int(metrics['total_vendedores'])}")
col3.metric("Valor Médio",  format_brl(metrics['valor_medio']))

st.markdown('---')

col_pie, col_bar = st.columns(2)

with col_pie:
    st.subheader('Distribuição por Vendedor')
    fig_pie = create_incentive_pie_chart(df_employee)
    st.plotly_chart(fig_pie, use_container_width=True)

with col_bar:
    st.subheader(f'Top {top_n_vendedores} Premiados')
    fig_bar = create_incentive_bar_chart(df_employee, top_n_vendedores)
    st.plotly_chart(fig_bar, use_container_width=True)

st.markdown('---')

todas_funcoes = extrair_funcoes_unicas(df_employee)

st.subheader('Filtrar por Função')

col_filtro, col_info_filtro = st.columns([3, 1])

with col_filtro:
    funcoes_selecionadas = st.multiselect(
        'Selecione as funções para filtrar a lista e a tabela mensal:',
        options=todas_funcoes,
        default=todas_funcoes,
        help="Afeta apenas a Lista de Consultores e a Comparação Mensal abaixo."
    )

with col_info_filtro:
    n_total    = len(df_employee)
    df_filtrado = filtrar_por_funcao(df_employee, funcoes_selecionadas)
    n_filtrado  = len(df_filtrado)
    st.metric("Premiados exibidos", f"{n_filtrado} / {n_total}")

ids_filtrados = df_filtrado['employee_id'].tolist()
df_monthly_filtrado = df_monthly[df_monthly['employee_id'].isin(ids_filtrados)]

if df_filtrado.empty:
    st.warning("Nenhum premiado encontrado para as funções selecionadas.")

st.markdown('---')

st.subheader('Lista de Consultores e Incentivos')

if not df_filtrado.empty and not df_monthly_filtrado.empty:
    vendedores_unicos = df_filtrado['vendedor'].tolist()
    
    for idx, vendedor in enumerate(vendedores_unicos):
        dados_vendedor = df_filtrado[df_filtrado['vendedor'] == vendedor].iloc[0]
        valor_total    = dados_vendedor['valor_total_incentivos']
        funcao         = dados_vendedor.get('funcao', 'N/A')
        
        dados_mensais = df_monthly_filtrado[
            df_monthly_filtrado['vendedor'] == vendedor
        ].sort_values('mes')
        
        with st.expander(
            f"**{vendedor}** — {format_brl(valor_total)} | {funcao}",
            expanded=False
        ):
            if not dados_mensais.empty:
                col_info, col_chart = st.columns([1, 2])
                
                with col_info:
                    st.markdown("**Detalhamento Mensal:**")
                    for _, row in dados_mensais.iterrows():
                        st.markdown(
                            f"- **{row['mes_display']}**: {format_brl(row['valor_mes'])} "
                            f"({int(row['quantidade_mes'])} incentivo(s))"
                        )
                
                with col_chart:
                    import plotly.graph_objects as go
                    fig_mini = go.Figure()
                    fig_mini.add_trace(go.Bar(
                        x=dados_mensais['mes_display'],
                        y=dados_mensais['valor_mes'],
                        text=dados_mensais['valor_mes'],
                        texttemplate='R$ %{text:,.2f}',
                        textposition='outside',
                        marker_color='#47C7DA'
                    ))
                    fig_mini.update_layout(
                        title=f"Evolução — {vendedor}",
                        xaxis_title="Mês",
                        yaxis_title="Valor (R$)",
                        height=300,
                        showlegend=False
                    )
                    st.plotly_chart(
                        fig_mini,
                        use_container_width=True,
                        key=f"chart_vendedor_{idx}"
                    )
            else:
                st.info("Sem detalhamento mensal disponível")
else:
    st.info("Sem dados para exibir com os filtros de função selecionados.")

st.markdown('---')


st.subheader('Comparação Mensal')

if not df_monthly_filtrado.empty:
    df_pivot = create_monthly_pivot_table(df_monthly_filtrado)
    
    if not df_pivot.empty:
        df_pivot_formatted = format_incentive_monthly_table(df_pivot)
        
        st.dataframe(
            df_pivot_formatted,
            use_container_width=True,
            height=400
        )
        
        from io import BytesIO
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_pivot.to_excel(writer, sheet_name='Incentivos Mensal', index=False)
            
            workbook  = writer.book
            worksheet = workbook['Incentivos Mensal']
            
            header_fill = PatternFill(start_color="47C7DA", end_color="47C7DA", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal='center')
            
            for row in worksheet.iter_rows(min_row=2, min_col=2):
                for cell in row:
                    if cell.column > 1:
                        cell.number_format = 'R$ #,##0.00'
                        cell.alignment     = Alignment(horizontal='right')
            
            for column in worksheet.columns:
                max_length    = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except Exception:
                        pass
                worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        excel_data = output.getvalue()
        
        funcoes_label = "_".join(funcoes_selecionadas[:2]) if funcoes_selecionadas else "todas"
        st.download_button(
            label="⬇️ Baixar Tabela Mensal (XLSX)",
            data=excel_data,
            file_name=(
                f"incentivos_{grupo_selecionado}_{funcoes_label}_"
                f"{datetime.now().strftime('%Y%m%d')}.xlsx"
            ),
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    else:
        st.info("Não foi possível criar tabela pivot")
else:
    st.info("Sem dados mensais para as funções selecionadas.")

st.markdown('---')

st.subheader('Comparação de Incentivos por Loja')

if not df_monthly.empty and not df_employee.empty:
    fig_lojas = create_monthly_stores_comparison_chart(df_monthly, df_employee)
    st.plotly_chart(fig_lojas, use_container_width=True)
    
    col_info1, col_info2 = st.columns(2)
    
    with col_info1:
        st.info(
            "**Gráfico mostra:**\n\n"
            "- Comparação entre lojas do grupo\n"
            "- Evolução mensal de incentivos\n"
            "- Total agregado (todas as funções)"
        )
    
    with col_info2:
        df_with_loja  = df_monthly.merge(
            df_employee[['vendedor', 'loja']],
            on='vendedor',
            how='left'
        )
        loja_destaque = df_with_loja.groupby('loja')['valor_mes'].sum().idxmax()
        valor_destaque = df_with_loja.groupby('loja')['valor_mes'].sum().max()
        
        st.metric(
            "Loja Destaque",
            loja_destaque,
            format_brl(valor_destaque)
        )
else:
    st.info("Sem dados para comparação de lojas")

st.markdown('---')
st.caption('Dados em tempo real do PostgreSQL')

#-------------------DEBUG--------------------
if st.secrets.get('settings', {}).get('debug_mode', False):
    with st.expander("Debug: Dados Brutos", expanded=False):
        st.write("### df_employee (completo)")
        st.dataframe(df_employee)
        
        st.write("### df_employee (filtrado por função)")
        st.dataframe(df_filtrado)
        
        st.write("### df_monthly (filtrado por função)")
        st.dataframe(df_monthly_filtrado)
        
        st.write("### Métricas Calculadas")
        st.json(metrics)
        
        st.write("### Funções disponíveis")
        st.write(todas_funcoes)
        
        st.write("### Funções selecionadas")
        st.write(funcoes_selecionadas)
        
        st.write("### Meses Selecionados")
        st.write(f"Display: {meses_selecionados_display}")
        st.write(f"Formato YYYY-MM: {meses_selecionados}")